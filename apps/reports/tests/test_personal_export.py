from io import BytesIO

import datetime as dt
import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.test import Client
from django.urls import reverse
from openpyxl import load_workbook

from apps.accounts.services import assign_buddy
from apps.learning.models import (
    Assessment,
    CapabilityCategory,
    CapabilityDomain,
    CapabilityItem,
    LearningCycle,
)
from apps.learning.services_execution import add_progress_update
from apps.learning.services_planning import approve_plan, generate_plan, submit_plan

User = get_user_model()


def _user(username, role):
    user = User.objects.create_user(
        username=username, password="testpass123", email=f"{username}@example.com"
    )
    user.groups.add(Group.objects.get(name=role))
    return user


def _capability():
    category = CapabilityCategory.objects.create(name="Tech", sort_order=1)
    l1 = CapabilityDomain.objects.create(
        category=category, code="T01", name="Backend", level=1, sort_order=1
    )
    l2 = CapabilityDomain.objects.create(
        category=category, parent=l1, code="T01.01", name="API", level=2, sort_order=1
    )
    return CapabilityItem.objects.create(
        domain=l2,
        code="T01.01.01",
        name="Contracts",
        suggested_level="P6",
        acceptance_method="Demo",
        estimated_hours="8",
        recommended_action="Build an API checklist",
        sort_order=1,
    )


@pytest.fixture
def member_plan(db):
    leader = _user("leader", "leader")
    member = _user("member", "member")
    buddy = _user("buddy", "buddy")
    assign_buddy(member, buddy)
    capability = _capability()
    cycle = LearningCycle.objects.create(
        cycle_type=LearningCycle.Type.CALENDAR_YEAR,
        year=2026,
        created_by=leader,
    )
    cycle.add_participant(member)
    assessment = Assessment.objects.get(
        cycle=cycle, member=member, capability_item=capability
    )
    assessment.current_level = 1
    assessment.target_level = 3
    assessment.priority = "high"
    assessment.included = True
    assessment.planned_quarter = "Q2"
    assessment.planned_month = dt.date(2026, 5, 1)
    assessment.save()
    plan = generate_plan(member, cycle)
    submit_plan(plan, member)
    approve_plan(plan, buddy)
    item = plan.items.get()
    add_progress_update(item, member, "Read and apply", "2.5")
    client = Client()
    client.force_login(member)
    return member, cycle, plan, client


@pytest.mark.django_db
def test_personal_export_contains_required_sheets(member_plan):
    _member, cycle, _plan, client = member_plan

    response = client.get(reverse("reports:personal", args=[cycle.pk]))

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["自评结果", "年度计划", "执行与验收"]


@pytest.mark.django_db
def test_personal_export_contains_assessment_plan_and_progress(member_plan):
    _member, cycle, _plan, client = member_plan

    response = client.get(reverse("reports:personal", args=[cycle.pk]))
    workbook = load_workbook(BytesIO(response.content), data_only=True)

    assert workbook["自评结果"]["A2"].value == "T01.01.01"
    assert workbook["年度计划"]["B2"].value == "Contracts"
    assert workbook["执行与验收"]["E2"].value == 2.5

