from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.test import Client
from django.urls import reverse
from openpyxl import load_workbook

from apps.learning.models import PlanItem

from .test_personal_export import member_plan

User = get_user_model()


@pytest.mark.django_db
def test_team_export_contains_summary_and_detail(member_plan):
    _member, cycle, plan, _client = member_plan
    leader = User.objects.get(username="leader")
    plan.items.update(execution_status=PlanItem.ExecutionStatus.COMPLETED)
    client = Client()
    client.force_login(leader)

    response = client.get(reverse("reports:team", args=[cycle.pk]))

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["团队汇总", "计划明细", "能力差距"]
    assert workbook["团队汇总"]["A2"].value == "member"
    assert workbook["团队汇总"]["D2"].value == 1
    assert workbook["计划明细"]["C2"].value == "T01.01.01"
    assert workbook["能力差距"]["A2"].value == "T01.01.01"


@pytest.mark.django_db
def test_member_cannot_download_team_export(member_plan):
    _member, cycle, _plan, client = member_plan

    response = client.get(reverse("reports:team", args=[cycle.pk]))

    assert response.status_code == 403

