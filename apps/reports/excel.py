from io import BytesIO

from openpyxl import Workbook

from apps.learning.models import Assessment, LearningPlan, PlanItem


def _clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return "".join(ch for ch in value if ch >= " " or ch in "\n\t")
    return value


def _append(ws, values):
    ws.append([_clean(value) for value in values])


def workbook_bytes(workbook):
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def personal_workbook(member, cycle):
    workbook = Workbook()
    assessment_ws = workbook.active
    assessment_ws.title = "自评结果"
    plan_ws = workbook.create_sheet("年度计划")
    execution_ws = workbook.create_sheet("执行与验收")

    _append(assessment_ws, ["能力编码", "能力名称", "当前", "目标", "差距", "优先级", "纳入计划"])
    for assessment in (
        Assessment.objects.filter(member=member, cycle=cycle)
        .select_related("capability_item")
        .order_by("capability_item__sort_order", "capability_item__code")
    ):
        _append(
            assessment_ws,
            [
                assessment.capability_item.code,
                assessment.capability_item.name,
                assessment.current_level,
                assessment.target_level,
                assessment.gap,
                assessment.priority,
                "是" if assessment.included else "否",
            ],
        )

    _append(plan_ws, ["月份", "能力名称", "任务", "验收方式", "预计耗时", "状态"])
    plan = (
        LearningPlan.objects.filter(member=member, cycle=cycle)
        .select_related("buddy")
        .first()
    )
    items = plan.items.all() if plan else PlanItem.objects.none()
    for item in items.order_by("planned_month", "sort_order", "capability_code"):
        _append(
            plan_ws,
            [
                item.planned_month,
                item.capability_name,
                item.task,
                item.acceptance_method,
                item.estimated_hours,
                item.get_execution_status_display(),
            ],
        )

    _append(execution_ws, ["能力编码", "能力名称", "进展", "提交说明", "实际耗时", "验收状态"])
    for item in items.prefetch_related("progress_updates", "evidence_submissions"):
        progress = "\n".join(update.content for update in item.progress_updates.all())
        evidence = "\n".join(submission.note for submission in item.evidence_submissions.all())
        _append(
            execution_ws,
            [
                item.capability_code,
                item.capability_name,
                progress,
                evidence,
                float(item.actual_hours),
                item.get_execution_status_display(),
            ],
        )
    return workbook


def team_workbook(cycle):
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "团队汇总"
    detail_ws = workbook.create_sheet("计划明细")
    gap_ws = workbook.create_sheet("能力差距")

    _append(summary_ws, ["成员", "Buddy", "计划项", "已完成", "延期", "实际耗时"])
    plans = (
        LearningPlan.objects.filter(cycle=cycle)
        .select_related("member", "buddy")
        .prefetch_related("items__progress_updates")
        .order_by("member__username")
    )
    for plan in plans:
        items = list(plan.items.all())
        completed = sum(
            1 for item in items if item.execution_status == PlanItem.ExecutionStatus.COMPLETED
        )
        hours = sum((item.actual_hours for item in items), 0)
        _append(
            summary_ws,
            [plan.member.username, plan.buddy.username, len(items), completed, 0, float(hours)],
        )

    _append(detail_ws, ["成员", "Buddy", "能力编码", "能力名称", "月份", "状态"])
    for plan in plans:
        for item in plan.items.order_by("planned_month", "sort_order", "capability_code"):
            _append(
                detail_ws,
                [
                    plan.member.username,
                    plan.buddy.username,
                    item.capability_code,
                    item.capability_name,
                    item.planned_month,
                    item.get_execution_status_display(),
                ],
            )

    _append(gap_ws, ["能力编码", "能力名称", "成员", "当前", "目标", "差距"])
    for assessment in (
        Assessment.objects.filter(cycle=cycle)
        .select_related("member", "capability_item")
        .order_by("capability_item__code", "member__username")
    ):
        _append(
            gap_ws,
            [
                assessment.capability_item.code,
                assessment.capability_item.name,
                assessment.member.username,
                assessment.current_level,
                assessment.target_level,
                assessment.gap,
            ],
        )
    return workbook

