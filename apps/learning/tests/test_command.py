import io
import os
import tempfile
from unittest.mock import patch

from django.core.management import call_command
from django.core.management.base import CommandError
from django.db import IntegrityError
from django.test import TestCase
from openpyxl import Workbook

from apps.learning.models import (
    CapabilityCategory,
    CapabilityDomain,
    CapabilityItem,
    CapabilityMaterial,
    LearningMaterial,
)


def _save_temp_workbook(wb):
    handle, path = tempfile.mkstemp(suffix=".xlsx")
    try:
        wb.save(path)
    finally:
        os.close(handle)
    return path


def _make_valid_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    mat_ws = wb.create_sheet("06_学习材料索引")
    mat_ws.append(["材料编码", "材料名称", "材料类型", "材料来源 / 附件名", "用途说明", "材料状态"])
    mat_ws.append(["P01-M001", "Product material", "Doc", "Internal", "Usage", "Active"])
    mat_ws.append(["P01-M002", "Deploy material", "Manual", "Internal", "Usage", "Active"])

    cap_ws = wb.create_sheet("02_能力差距自评")
    cap_ws.append([
        "能力类别", "一级序号", "一级能力域", "二级序号", "二级能力项",
        "三级序号", "三级能力项", "建议起始职级", "学习材料",
        "预期输出 / 验收方式", "预计耗时(h)", "当前掌握度(0-5)",
        "目标掌握度(0-5)", "差距", "优先级", "是否纳入计划", "计划季度",
        "计划完成月份", "完成状态", "实际输出链接/说明", "备注", "推荐行动",
        "入选顺序(自动)",
    ])
    cap_ws.append([
        "专业能力", "P01", "Data Infra", "P01.01", "Cognition",
        "P01.01.01", "Product positioning", "P4", "P01-M001",
        "Screenshot", "4–6", None, None, None, None, None, None, None,
        "未开始", None, None, None, None, None,
    ])
    cap_ws.append([
        "专业能力", "P01", "Data Infra", "P01.01", "Cognition",
        "P01.01.02", "Boundary", "P5", "P01-M001、P01-M002",
        "Mapping table", "4–8", None, None, None, None, None, None, None,
        "未开始", None, None, None, None, None,
    ])

    return _save_temp_workbook(wb)


class ImportCommandTests(TestCase):
    def test_dry_run_does_not_write(self):
        path = _make_valid_workbook()
        out = io.StringIO()
        call_command("import_learning_template", path, stdout=out)

        output = out.getvalue()
        self.assertEqual(CapabilityCategory.objects.count(), 0)
        self.assertEqual(CapabilityDomain.objects.count(), 0)
        self.assertEqual(CapabilityItem.objects.count(), 0)
        self.assertEqual(LearningMaterial.objects.count(), 0)
        self.assertIn("categories", output)
        self.assertIn("materials", output)

    def test_apply_creates_records(self):
        path = _make_valid_workbook()
        out = io.StringIO()
        call_command("import_learning_template", path, "--apply", stdout=out)

        self.assertEqual(CapabilityCategory.objects.count(), 1)
        self.assertEqual(CapabilityDomain.objects.count(), 2)
        self.assertEqual(CapabilityItem.objects.count(), 2)
        self.assertEqual(LearningMaterial.objects.count(), 2)
        self.assertEqual(CapabilityMaterial.objects.count(), 3)

        item = CapabilityItem.objects.get(code="P01.01.02")
        self.assertEqual(item.materials.count(), 2)
        self.assertIn("Imported", out.getvalue())

    def test_apply_refuses_when_catalog_populated(self):
        CapabilityCategory.objects.create(name="Existing", sort_order=0)
        path = _make_valid_workbook()

        with self.assertRaises(CommandError) as cm:
            call_command("import_learning_template", path, "--apply")
        self.assertIn("already", str(cm.exception).lower())

    def test_apply_handles_integrity_error(self):
        path = _make_valid_workbook()

        with patch.object(
            CapabilityDomain.objects, "create", side_effect=IntegrityError("duplicate")
        ):
            with self.assertRaises(CommandError) as cm:
                call_command("import_learning_template", path, "--apply")

        self.assertIn("concurrently", str(cm.exception).lower())
        self.assertEqual(CapabilityCategory.objects.count(), 0)
        self.assertEqual(CapabilityDomain.objects.count(), 0)
        self.assertEqual(CapabilityItem.objects.count(), 0)
        self.assertEqual(LearningMaterial.objects.count(), 0)
        self.assertEqual(CapabilityMaterial.objects.count(), 0)

    def test_apply_rolls_back_on_invalid_workbook(self):
        wb = Workbook()
        wb.remove(wb.active)
        mat_ws = wb.create_sheet("06_学习材料索引")
        mat_ws.append(["材料编码", "材料名称", "材料类型", "材料来源 / 附件名", "用途说明", "材料状态"])
        mat_ws.append(["P01-M001", "Material", "Doc", "Internal", "Usage", "Active"])

        cap_ws = wb.create_sheet("02_能力差距自评")
        cap_ws.append([
            "能力类别", "一级序号", "一级能力域", "二级序号", "二级能力项",
            "三级序号", "三级能力项", "建议起始职级", "学习材料",
            "预期输出 / 验收方式", "预计耗时(h)", "当前掌握度(0-5)",
            "目标掌握度(0-5)", "差距", "优先级", "是否纳入计划", "计划季度",
            "计划完成月份", "完成状态", "实际输出链接/说明", "备注", "推荐行动",
            "入选顺序(自动)",
        ])
        cap_ws.append([
            "专业能力", "P01", "Data Infra", "P01.01", "Cognition",
            "P01.01.01", "First", "P4", "P01-M999", "Output", "4", None, None,
            None, None, None, None, None, "未开始", None, None, None, None, None,
        ])

        path = _save_temp_workbook(wb)

        with self.assertRaises(CommandError):
            call_command("import_learning_template", path, "--apply")

        self.assertEqual(CapabilityCategory.objects.count(), 0)
        self.assertEqual(LearningMaterial.objects.count(), 0)
        self.assertEqual(CapabilityItem.objects.count(), 0)
