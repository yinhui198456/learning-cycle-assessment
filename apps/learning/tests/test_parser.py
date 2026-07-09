import os
import re
import tempfile
from pathlib import Path

from django.test import TestCase
from openpyxl import Workbook

from apps.learning.parser import LearningTemplateError, parse_learning_template


def _save_temp_workbook(wb):
    handle, path = tempfile.mkstemp(suffix=".xlsx")
    try:
        wb.save(path)
    finally:
        os.close(handle)
    return path


def _make_valid_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    mat_ws = wb.create_sheet("06_学习材料索引")
    mat_ws.append(["材料编码", "材料名称", "材料类型", "材料来源 / 附件名", "用途说明", "材料状态"])
    mat_ws.append(["P01-M001", "Product material", "Doc", "Internal", "Usage", "Active"])
    mat_ws.append(["P01-M002", "Deploy material", "Manual", "Internal", "Usage", "Active"])
    mat_ws.append(["C01-M001", "General material", "Video", "Web", "Usage", "Active"])

    cap_ws = wb.create_sheet("02_能力差距自评")
    cap_ws.append([
        "能力类别", "一级序号", "一级能力域", "二级序号", "二级能力项",
        "三级序号", "三级能力项", "建议起始职级", "学习材料",
        "预期输出 / 验收方式", "预计耗时(h)", "当前掌握度(0-5)",
        "目标掌握度(0-5)", "差距", "优先级", "是否纳入计划", "计划季度",
        "计划完成月份", "完成状态", "实际输出链接/说明", "备注", "推荐行动",
        "入选顺序(自动)",
    ])
    cap_ws.append([
        "专业能力", "P01", "Data Infra", "P01.01", "Cognition",
        "P01.01.01", "Product positioning", "P4", "P01-M001、内部文档",
        "Screenshot", "4–6", None, None, None, None, None, None, None,
        "未开始", None, None, None, None, None,
    ])
    cap_ws.append([
        "专业能力", "P01", "Data Infra", "P01.01", "Cognition",
        "P01.01.02", "Boundary", "P5", "P01-M001、P01-M002、内部图示样例",
        "Mapping table", "4–8", None, None, None, None, None, None, None,
        "未开始", None, None, None, None, None,
    ])
    cap_ws.append([
        "通用素质能力", "C01", "Communication", "C01.01", "Expression",
        "C01.01.01", "Active listening", "P4", "C01-M001",
        "Report", "2", None, None, None, None, None, None, None,
        "未开始", None, None, "推荐行动A", None, None,
    ])

    return _save_temp_workbook(wb)


class ParserTests(TestCase):
    def test_valid_workbook_parses(self):
        path = _make_valid_workbook()
        result = parse_learning_template(path)

        self.assertEqual(result["categories"], [
            {"name": "专业能力", "sort_order": 0, "is_active": True},
            {"name": "通用素质能力", "sort_order": 1, "is_active": True},
        ])
        self.assertEqual(len(result["materials"]), 3)
        self.assertEqual(len(result["domains"]), 4)
        self.assertEqual(len(result["items"]), 3)

        domain_codes = {d["code"] for d in result["domains"]}
        self.assertEqual(domain_codes, {"P01", "P01.01", "C01", "C01.01"})

        item = next(i for i in result["items"] if i["code"] == "P01.01.02")
        self.assertEqual(item["name"], "Boundary")
        self.assertEqual(item["suggested_level"], "P5")
        self.assertEqual(item["material_reference"], "P01-M001、P01-M002、内部图示样例")
        self.assertEqual(item["material_codes"], ["P01-M001", "P01-M002"])
        self.assertEqual(item["acceptance_method"], "Mapping table")
        self.assertEqual(item["estimated_hours"], "4–8")
        self.assertEqual(item["recommended_action"], "")

    def test_mixed_material_reference_keeps_free_text(self):
        path = _make_valid_workbook()
        result = parse_learning_template(path)

        item = next(i for i in result["items"] if i["code"] == "P01.01.01")
        self.assertEqual(item["material_reference"], "P01-M001、内部文档")
        self.assertEqual(item["material_codes"], ["P01-M001"])

    def test_items_resolve_to_level_two_domains(self):
        path = _make_valid_workbook()
        result = parse_learning_template(path)

        domain_by_code = {d["code"]: d for d in result["domains"]}
        for item in result["items"]:
            domain = domain_by_code[item["domain_code"]]
            self.assertEqual(domain["level"], 2)

    def test_missing_capability_sheet_raises(self):
        wb = Workbook()
        wb.remove(wb.active)
        mat_ws = wb.create_sheet("06_学习材料索引")
        mat_ws.append(["材料编码", "材料名称", "材料类型", "材料来源 / 附件名", "用途说明", "材料状态"])
        mat_ws.append(["P01-M001", "Material", "Doc", "Internal", "Usage", "Active"])

        path = _save_temp_workbook(wb)

        with self.assertRaises(LearningTemplateError) as cm:
            parse_learning_template(path)
        self.assertIn("02_能力差距自评", str(cm.exception))

    def test_missing_material_sheet_raises(self):
        wb = Workbook()
        wb.remove(wb.active)
        cap_ws = wb.create_sheet("02_能力差距自评")
        cap_ws.append(["能力类别"])

        path = _save_temp_workbook(wb)

        with self.assertRaises(LearningTemplateError) as cm:
            parse_learning_template(path)
        self.assertIn("06_学习材料索引", str(cm.exception))

    def test_missing_header_raises(self):
        wb = Workbook()
        wb.remove(wb.active)
        mat_ws = wb.create_sheet("06_学习材料索引")
        mat_ws.append(["Wrong header"])
        cap_ws = wb.create_sheet("02_能力差距自评")
        cap_ws.append(["能力类别"])

        path = _save_temp_workbook(wb)

        with self.assertRaises(LearningTemplateError) as cm:
            parse_learning_template(path)
        self.assertIn("header", str(cm.exception).lower())

    def test_duplicate_item_code_raises(self):
        wb = Workbook()
        wb.remove(wb.active)
        mat_ws = wb.create_sheet("06_学习材料索引")
        mat_ws.append(["材料编码", "材料名称", "材料类型", "材料来源 / 附件名", "用途说明", "材料状态"])
        mat_ws.append(["P01-M001", "Material", "Doc", "Internal", "Usage", "Active"])

        cap_ws = wb.create_sheet("02_能力差距自评")
        cap_ws.append([
            "能力类别", "一级序号", "一级能力域", "二级序号", "二级能力项",
            "三级序号", "三级能力项", "建议起始职级", "学习材料",
            "预期输出 / 验收方式", "预计耗时(h)", "当前掌握度(0-5)",
            "目标掌握度(0-5)", "差距", "优先级", "是否纳入计划", "计划季度",
            "计划完成月份", "完成状态", "实际输出链接/说明", "备注", "推荐行动",
            "入选顺序(自动)",
        ])
        cap_ws.append([
            "专业能力", "P01", "Data Infra", "P01.01", "Cognition",
            "P01.01.01", "First", "P4", "P01-M001", "Output", "4", None, None,
            None, None, None, None, None, "未开始", None, None, None, None, None,
        ])
        cap_ws.append([
            "专业能力", "P01", "Data Infra", "P01.01", "Cognition",
            "P01.01.01", "Second", "P5", "P01-M001", "Output", "6", None, None,
            None, None, None, None, None, "未开始", None, None, None, None, None,
        ])

        path = _save_temp_workbook(wb)

        with self.assertRaises(LearningTemplateError) as cm:
            parse_learning_template(path)
        self.assertIn("P01.01.01", str(cm.exception))

    def test_conflicting_domain_raises(self):
        wb = Workbook()
        wb.remove(wb.active)
        mat_ws = wb.create_sheet("06_学习材料索引")
        mat_ws.append(["材料编码", "材料名称", "材料类型", "材料来源 / 附件名", "用途说明", "材料状态"])
        mat_ws.append(["P01-M001", "Material", "Doc", "Internal", "Usage", "Active"])

        cap_ws = wb.create_sheet("02_能力差距自评")
        cap_ws.append([
            "能力类别", "一级序号", "一级能力域", "二级序号", "二级能力项",
            "三级序号", "三级能力项", "建议起始职级", "学习材料",
            "预期输出 / 验收方式", "预计耗时(h)", "当前掌握度(0-5)",
            "目标掌握度(0-5)", "差距", "优先级", "是否纳入计划", "计划季度",
            "计划完成月份", "完成状态", "实际输出链接/说明", "备注", "推荐行动",
            "入选顺序(自动)",
        ])
        cap_ws.append([
            "专业能力", "P01", "Data Infra", "P01.01", "Cognition",
            "P01.01.01", "First", "P4", "P01-M001", "Output", "4", None, None,
            None, None, None, None, None, "未开始", None, None, None, None, None,
        ])
        cap_ws.append([
            "通用素质能力", "P01", "Other Infra", "P01.02", "Other",
            "P01.02.01", "Second", "P5", "P01-M001", "Output", "6", None, None,
            None, None, None, None, None, "未开始", None, None, None, None, None,
        ])

        path = _save_temp_workbook(wb)

        with self.assertRaises(LearningTemplateError) as cm:
            parse_learning_template(path)
        self.assertIn("P01", str(cm.exception))

    def test_level1_code_reused_as_level2_raises(self):
        wb = Workbook()
        wb.remove(wb.active)
        mat_ws = wb.create_sheet("06_学习材料索引")
        mat_ws.append(["材料编码", "材料名称", "材料类型", "材料来源 / 附件名", "用途说明", "材料状态"])
        mat_ws.append(["P01-M001", "Material", "Doc", "Internal", "Usage", "Active"])

        cap_ws = wb.create_sheet("02_能力差距自评")
        cap_ws.append([
            "能力类别", "一级序号", "一级能力域", "二级序号", "二级能力项",
            "三级序号", "三级能力项", "建议起始职级", "学习材料",
            "预期输出 / 验收方式", "预计耗时(h)", "当前掌握度(0-5)",
            "目标掌握度(0-5)", "差距", "优先级", "是否纳入计划", "计划季度",
            "计划完成月份", "完成状态", "实际输出链接/说明", "备注", "推荐行动",
            "入选顺序(自动)",
        ])
        cap_ws.append([
            "专业能力", "P01", "Data Infra", "P01.01", "Cognition",
            "P01.01.01", "First", "P4", "P01-M001", "Output", "4", None, None,
            None, None, None, None, None, "未开始", None, None, None, None, None,
        ])
        cap_ws.append([
            "专业能力", "P01.01", "Cognition", "P01.02", "Other",
            "P01.02.01", "Second", "P5", "P01-M001", "Output", "6", None, None,
            None, None, None, None, None, "未开始", None, None, None, None, None,
        ])

        path = _save_temp_workbook(wb)

        with self.assertRaises(LearningTemplateError) as cm:
            parse_learning_template(path)
        self.assertIn("P01.01", str(cm.exception))

    def test_level2_code_reused_as_level1_raises(self):
        wb = Workbook()
        wb.remove(wb.active)
        mat_ws = wb.create_sheet("06_学习材料索引")
        mat_ws.append(["材料编码", "材料名称", "材料类型", "材料来源 / 附件名", "用途说明", "材料状态"])
        mat_ws.append(["P01-M001", "Material", "Doc", "Internal", "Usage", "Active"])

        cap_ws = wb.create_sheet("02_能力差距自评")
        cap_ws.append([
            "能力类别", "一级序号", "一级能力域", "二级序号", "二级能力项",
            "三级序号", "三级能力项", "建议起始职级", "学习材料",
            "预期输出 / 验收方式", "预计耗时(h)", "当前掌握度(0-5)",
            "目标掌握度(0-5)", "差距", "优先级", "是否纳入计划", "计划季度",
            "计划完成月份", "完成状态", "实际输出链接/说明", "备注", "推荐行动",
            "入选顺序(自动)",
        ])
        cap_ws.append([
            "专业能力", "P01.01", "Data", "P01", "Cognition",
            "P01.01.01", "First", "P4", "P01-M001", "Output", "4", None, None,
            None, None, None, None, None, "未开始", None, None, None, None, None,
        ])
        cap_ws.append([
            "专业能力", "P01", "Cognition", "P01.02", "Other",
            "P01.02.01", "Second", "P5", "P01-M001", "Output", "6", None, None,
            None, None, None, None, None, "未开始", None, None, None, None, None,
        ])

        path = _save_temp_workbook(wb)

        with self.assertRaises(LearningTemplateError) as cm:
            parse_learning_template(path)
        self.assertIn("P01", str(cm.exception))

    def test_hierarchy_mismatch_raises(self):
        wb = Workbook()
        wb.remove(wb.active)
        mat_ws = wb.create_sheet("06_学习材料索引")
        mat_ws.append(["材料编码", "材料名称", "材料类型", "材料来源 / 附件名", "用途说明", "材料状态"])
        mat_ws.append(["P01-M001", "Material", "Doc", "Internal", "Usage", "Active"])

        cap_ws = wb.create_sheet("02_能力差距自评")
        cap_ws.append([
            "能力类别", "一级序号", "一级能力域", "二级序号", "二级能力项",
            "三级序号", "三级能力项", "建议起始职级", "学习材料",
            "预期输出 / 验收方式", "预计耗时(h)", "当前掌握度(0-5)",
            "目标掌握度(0-5)", "差距", "优先级", "是否纳入计划", "计划季度",
            "计划完成月份", "完成状态", "实际输出链接/说明", "备注", "推荐行动",
            "入选顺序(自动)",
        ])
        cap_ws.append([
            "专业能力", "P01", "Data Infra", "P01.01", "Cognition",
            "P01.01.01", "First", "P4", "P01-M001", "Output", "4", None, None,
            None, None, None, None, None, "未开始", None, None, None, None, None,
        ])
        cap_ws.append([
            "专业能力", "P02", "Other Domain", "P01.01", "Cognition",
            "P01.01.02", "Second", "P5", "P01-M001", "Output", "6", None, None,
            None, None, None, None, None, "未开始", None, None, None, None, None,
        ])

        path = _save_temp_workbook(wb)

        with self.assertRaises(LearningTemplateError) as cm:
            parse_learning_template(path)
        self.assertIn("P01.01", str(cm.exception))

    def test_unknown_code_shaped_material_reference_raises(self):
        wb = Workbook()
        wb.remove(wb.active)
        mat_ws = wb.create_sheet("06_学习材料索引")
        mat_ws.append(["材料编码", "材料名称", "材料类型", "材料来源 / 附件名", "用途说明", "材料状态"])
        mat_ws.append(["P01-M001", "Material", "Doc", "Internal", "Usage", "Active"])

        cap_ws = wb.create_sheet("02_能力差距自评")
        cap_ws.append([
            "能力类别", "一级序号", "一级能力域", "二级序号", "二级能力项",
            "三级序号", "三级能力项", "建议起始职级", "学习材料",
            "预期输出 / 验收方式", "预计耗时(h)", "当前掌握度(0-5)",
            "目标掌握度(0-5)", "差距", "优先级", "是否纳入计划", "计划季度",
            "计划完成月份", "完成状态", "实际输出链接/说明", "备注", "推荐行动",
            "入选顺序(自动)",
        ])
        cap_ws.append([
            "专业能力", "P01", "Data Infra", "P01.01", "Cognition",
            "P01.01.01", "First", "P4", "P01-M999、内部文档", "Output", "4",
            None, None, None, None, None, None, None, "未开始", None, None, None,
            None, None,
        ])

        path = _save_temp_workbook(wb)

        with self.assertRaises(LearningTemplateError) as cm:
            parse_learning_template(path)
        self.assertIn("P01-M999", str(cm.exception))

    def test_exact_duplicate_item_code_raises(self):
        wb = Workbook()
        wb.remove(wb.active)
        mat_ws = wb.create_sheet("06_学习材料索引")
        mat_ws.append(["材料编码", "材料名称", "材料类型", "材料来源 / 附件名", "用途说明", "材料状态"])
        mat_ws.append(["P01-M001", "Material", "Doc", "Internal", "Usage", "Active"])

        cap_ws = wb.create_sheet("02_能力差距自评")
        cap_ws.append([
            "能力类别", "一级序号", "一级能力域", "二级序号", "二级能力项",
            "三级序号", "三级能力项", "建议起始职级", "学习材料",
            "预期输出 / 验收方式", "预计耗时(h)", "当前掌握度(0-5)",
            "目标掌握度(0-5)", "差距", "优先级", "是否纳入计划", "计划季度",
            "计划完成月份", "完成状态", "实际输出链接/说明", "备注", "推荐行动",
            "入选顺序(自动)",
        ])
        cap_ws.append([
            "专业能力", "P01", "Data Infra", "P01.01", "Cognition",
            "P01.01.01", "Same", "P4", "P01-M001", "Output", "4", None, None,
            None, None, None, None, None, "未开始", None, None, None, None, None,
        ])
        cap_ws.append([
            "专业能力", "P01", "Data Infra", "P01.01", "Cognition",
            "P01.01.01", "Same", "P4", "P01-M001", "Output", "4", None, None,
            None, None, None, None, None, "未开始", None, None, None, None, None,
        ])

        path = _save_temp_workbook(wb)

        with self.assertRaises(LearningTemplateError) as cm:
            parse_learning_template(path)
        self.assertIn("P01.01.01", str(cm.exception))

    def test_repeated_material_reference_deduplicates_codes_preserves_raw_text(self):
        wb = Workbook()
        wb.remove(wb.active)
        mat_ws = wb.create_sheet("06_学习材料索引")
        mat_ws.append(["材料编码", "材料名称", "材料类型", "材料来源 / 附件名", "用途说明", "材料状态"])
        mat_ws.append(["P01-M001", "Material", "Doc", "Internal", "Usage", "Active"])

        cap_ws = wb.create_sheet("02_能力差距自评")
        cap_ws.append([
            "能力类别", "一级序号", "一级能力域", "二级序号", "二级能力项",
            "三级序号", "三级能力项", "建议起始职级", "学习材料",
            "预期输出 / 验收方式", "预计耗时(h)", "当前掌握度(0-5)",
            "目标掌握度(0-5)", "差距", "优先级", "是否纳入计划", "计划季度",
            "计划完成月份", "完成状态", "实际输出链接/说明", "备注", "推荐行动",
            "入选顺序(自动)",
        ])
        cap_ws.append([
            "专业能力", "P01", "Data Infra", "P01.01", "Cognition",
            "P01.01.01", "First", "P4", "P01-M001、P01-M001、内部文档",
            "Output", "4", None, None, None, None, None, None, None,
            "未开始", None, None, None, None, None,
        ])

        path = _save_temp_workbook(wb)
        result = parse_learning_template(path)

        item = next(i for i in result["items"] if i["code"] == "P01.01.01")
        self.assertEqual(item["material_reference"], "P01-M001、P01-M001、内部文档")
        self.assertEqual(item["material_codes"], ["P01-M001"])

    def test_duplicate_material_code_raises(self):
        wb = Workbook()
        wb.remove(wb.active)
        mat_ws = wb.create_sheet("06_学习材料索引")
        mat_ws.append(["材料编码", "材料名称", "材料类型", "材料来源 / 附件名", "用途说明", "材料状态"])
        mat_ws.append(["P01-M001", "First", "Doc", "Internal", "Usage", "Active"])
        mat_ws.append(["P01-M001", "Second", "Manual", "Web", "Usage", "Active"])

        cap_ws = wb.create_sheet("02_能力差距自评")
        cap_ws.append([
            "能力类别", "一级序号", "一级能力域", "二级序号", "二级能力项",
            "三级序号", "三级能力项", "建议起始职级", "学习材料",
            "预期输出 / 验收方式", "预计耗时(h)", "当前掌握度(0-5)",
            "目标掌握度(0-5)", "差距", "优先级", "是否纳入计划", "计划季度",
            "计划完成月份", "完成状态", "实际输出链接/说明", "备注", "推荐行动",
            "入选顺序(自动)",
        ])
        cap_ws.append([
            "专业能力", "P01", "Data Infra", "P01.01", "Cognition",
            "P01.01.01", "First", "P4", "P01-M001", "Output", "4", None, None,
            None, None, None, None, None, "未开始", None, None, None, None, None,
        ])

        path = _save_temp_workbook(wb)

        with self.assertRaises(LearningTemplateError) as cm:
            parse_learning_template(path)
        self.assertIn("P01-M001", str(cm.exception))
