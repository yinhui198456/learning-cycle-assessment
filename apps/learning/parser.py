import re
from pathlib import Path

from openpyxl import load_workbook


class LearningTemplateError(Exception):
    pass


_MATERIAL_SHEET = "06_学习材料索引"
_CAPABILITY_SHEET = "02_能力差距自评"

_MATERIAL_HEADER = (
    "材料编码",
    "材料名称",
    "材料类型",
    "材料来源 / 附件名",
    "用途说明",
    "材料状态",
)

_CAPABILITY_HEADER = (
    "能力类别",
    "一级序号",
    "一级能力域",
    "二级序号",
    "二级能力项",
    "三级序号",
    "三级能力项",
    "建议起始职级",
    "学习材料",
    "预期输出 / 验收方式",
    "预计耗时(h)",
    "当前掌握度(0-5)",
    "目标掌握度(0-5)",
    "差距",
    "优先级",
    "是否纳入计划",
    "计划季度",
    "计划完成月份",
    "完成状态",
    "实际输出链接/说明",
    "备注",
    "推荐行动",
    "入选顺序(自动)",
)

_MATERIAL_CODE_RE = re.compile(r"^[PC]\d{2}-M\d{3}$")


def _norm(value):
    if value is None:
        return ""
    return str(value).strip()


def _is_empty_row(row):
    return all(_norm(cell) == "" for cell in row)


def _assert_sheet(wb, name):
    if name not in wb.sheetnames:
        raise LearningTemplateError(f"Missing required sheet: {name}")
    return wb[name]


def _assert_header(ws, expected):
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise LearningTemplateError(f"Empty sheet: {ws.title}")
    actual = tuple(_norm(cell) for cell in header)
    while actual and actual[-1] == "":
        actual = actual[:-1]
    if actual != expected:
        raise LearningTemplateError(
            f"Unexpected header in sheet {ws.title}: {actual!r}"
        )
    return rows


def _parse_materials(rows):
    materials = []
    by_code = {}
    for idx, row in enumerate(rows, start=2):
        if _is_empty_row(row):
            continue
        code = _norm(row[0])
        name = _norm(row[1])
        if not code or not name:
            raise LearningTemplateError(
                f"Material row {idx} is missing code or name"
            )
        if code in by_code:
            raise LearningTemplateError(f"Duplicate material code: {code}")
        material = {
            "code": code,
            "name": name,
            "material_type": _norm(row[2]),
            "source": _norm(row[3]),
            "description": _norm(row[4]),
            "status": _norm(row[5]),
            "sort_order": len(materials),
        }
        materials.append(material)
        by_code[code] = material
    return materials, by_code


def _parse_capabilities(rows, materials_by_code):
    categories = {}
    category_order = []
    domains = []
    domain_by_code = {}
    items = []
    item_by_code = {}

    for idx, row in enumerate(rows, start=2):
        if _is_empty_row(row):
            continue
        category_name = _norm(row[0])
        l1_code = _norm(row[1])
        l1_name = _norm(row[2])
        l2_code = _norm(row[3])
        l2_name = _norm(row[4])
        item_code = _norm(row[5])
        item_name = _norm(row[6])
        suggested_level = _norm(row[7])
        material_reference = _norm(row[8])
        acceptance_method = _norm(row[9])
        estimated_hours = _norm(row[10])
        recommended_action = _norm(row[21])

        if not all(
            [category_name, l1_code, l1_name, l2_code, l2_name, item_code, item_name]
        ):
            raise LearningTemplateError(
                f"Capability row {idx} is missing required hierarchy fields"
            )

        if category_name not in categories:
            categories[category_name] = {
                "name": category_name,
                "sort_order": len(categories),
                "is_active": True,
            }
            category_order.append(category_name)

        existing_l1 = domain_by_code.get(l1_code)
        if existing_l1 is not None:
            if existing_l1["level"] != 1:
                raise LearningTemplateError(
                    f"Domain code {l1_code} is already registered as level {existing_l1['level']}"
                )
            if existing_l1["name"] != l1_name or existing_l1["category_name"] != category_name:
                raise LearningTemplateError(
                    f"Conflicting data for level-1 domain {l1_code}"
                )
        else:
            domain = {
                "category_name": category_name,
                "parent_code": None,
                "code": l1_code,
                "name": l1_name,
                "level": 1,
                "sort_order": len(domains),
            }
            domains.append(domain)
            domain_by_code[l1_code] = domain

        existing_l2 = domain_by_code.get(l2_code)
        if existing_l2 is not None:
            if existing_l2["level"] != 2:
                raise LearningTemplateError(
                    f"Domain code {l2_code} is already registered as level {existing_l2['level']}"
                )
            if existing_l2["name"] != l2_name or existing_l2["parent_code"] != l1_code:
                raise LearningTemplateError(
                    f"Level-2 domain {l2_code} does not belong to level-1 {l1_code}"
                )
            if existing_l2["category_name"] != category_name:
                raise LearningTemplateError(
                    f"Conflicting category for level-2 domain {l2_code}"
                )
        else:
            domain = {
                "category_name": category_name,
                "parent_code": l1_code,
                "code": l2_code,
                "name": l2_name,
                "level": 2,
                "sort_order": len(domains),
            }
            domains.append(domain)
            domain_by_code[l2_code] = domain

        existing_item = item_by_code.get(item_code)
        if existing_item is not None:
            raise LearningTemplateError(f"Duplicate capability item code: {item_code}")

        material_codes = []
        seen_material_codes = set()
        for token in material_reference.split("、"):
            token = token.strip()
            if not token:
                continue
            if _MATERIAL_CODE_RE.fullmatch(token):
                if token not in materials_by_code:
                    raise LearningTemplateError(
                        f"Unknown material reference: {token}"
                    )
                if token not in seen_material_codes:
                    material_codes.append(token)
                    seen_material_codes.add(token)

        item = {
            "domain_code": l2_code,
            "code": item_code,
            "name": item_name,
            "suggested_level": suggested_level,
            "material_reference": material_reference,
            "acceptance_method": acceptance_method,
            "estimated_hours": estimated_hours,
            "recommended_action": recommended_action,
            "sort_order": len(items),
            "material_codes": material_codes,
        }
        items.append(item)
        item_by_code[item_code] = item

    for item in items:
        domain = domain_by_code.get(item["domain_code"])
        if domain is None or domain["level"] != 2:
            raise LearningTemplateError(
                f"Capability item {item['code']} does not resolve to a level-2 domain"
            )

    return {
        "categories": [categories[name] for name in category_order],
        "domains": domains,
        "items": items,
    }


def parse_learning_template(path):
    path = Path(path)
    if not path.exists():
        raise LearningTemplateError(f"File not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        material_ws = _assert_sheet(wb, _MATERIAL_SHEET)
        material_rows = _assert_header(material_ws, _MATERIAL_HEADER)
        materials, materials_by_code = _parse_materials(material_rows)

        capability_ws = _assert_sheet(wb, _CAPABILITY_SHEET)
        capability_rows = _assert_header(capability_ws, _CAPABILITY_HEADER)
        parsed = _parse_capabilities(capability_rows, materials_by_code)

        return {
            "categories": parsed["categories"],
            "materials": materials,
            "domains": parsed["domains"],
            "items": parsed["items"],
        }
    finally:
        wb.close()
